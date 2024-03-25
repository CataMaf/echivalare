import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox
import pandas as pd
import os
import unicodedata
from openpyxl import load_workbook

# selectare fisier csv cu datele privind situatia scolara si generarea unui dataframe, extragere nume student si numar matricol
def select_csv_file(): 
    try:   
        global csv_df
        global csv_file_path
        csv_file_path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
        csv_df = pd.read_csv(csv_file_path)   
        get_student_name()
        get_matricola()
    except UnicodeDecodeError:
        messagebox.showerror("Eroare","Selectati fisierul care contine situatia scolara in format .csv. Ati selectat alt format de fisier.")
    
#selectarea fisierului excel cu macheta procesului verbal de echivalare
def select_excel_file():      
    global pv_path
    pv_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])        
   
#selectare fisier excel cu macheta catalogului de diferente
def select_catalog_diferente():
    global catalog_path
    catalog_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])

# obtinere nume student pentru a fi folosit in denumirea folderului si a numelui fisierului generat 
def get_student_name():       
    global student_name
    student_name = csv_df.at[0,'textbox10']
    
# obtinere numar matricol
def get_matricola():
    global nr_matricol    
    matricola = csv_df.at[0, 'textbox9']    
    mm = matricola.split('nr. ',1)
    mm1 = mm[1].split(', ',1)
    nr_matricol= mm1[0]

#normalizarea unui sir de caractere
# - pentru a putea normaliza denumirile materiilor pentrua a putea fi comparate 
#(normalizarea caracterelor si eliminare diacrtice)
def normalize_string(string):
    if type(string) is not str:
        return string
    string = string.strip()
    string = unicodedata.normalize('NFD', string)
    string =''.join(c for c in string if not unicodedata.combining(c))
    return string   

# extragem din fisierul csv datele referitoare la exmatriculare sau intrrupere studii
def get_exmatriculare_intrerupere():
    with open(csv_file_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()
        for i in range(len(lines)-1):
            if lines[i].strip() == '':
                if 'year' in lines[i-1].strip().lower():
                    return lines[i-2]                    
                else:
                    return lines[i-1]
    return ''
                    

#prelucrare dataframe obtinut din fisierul csv
def prelucrare_csv(dataframe):
    
    # stergerea datelor care nu sunt necesare de la inceputul fisierului
    dataframe.drop(index = dataframe.index[:2], axis=0, inplace =True)

    #redenumirea coloanei care contine denumirea materiilor
    dataframe.rename(columns = {'textbox7':'Materia'}, inplace = True)

    #inlocuirea caractere nenumerice pentru a putea face calcule
    dataframe.replace('Abs',0, inplace=True)
    dataframe.replace('Adm',11,inplace=True)
    dataframe['textbox10'].fillna(0, inplace=True)
    dataframe['textbox11'].fillna(0,inplace=True)
    dataframe['textbox12'].fillna(0, inplace=True) 

    #stergere randuri care contine altfel de date decat cele relevante 
    indexes = dataframe.index[dataframe['textbox10'].str.contains('textbox', na=False)].tolist()
    for ind in indexes:
        dataframe.drop(index=ind, inplace=True)  
    
    # adaugarea unei coloane noi "Nota" ca suma a celor trei coloane initiale care ar putea contine note
    dataframe['Nota']= dataframe['textbox10'].apply(int) + dataframe['textbox11'].apply(int)+dataframe['textbox12'].apply(int)

    # normalizam denumirea materiilor pentru a putea fi comparate cu materiile din celalalt fisier
    list = dataframe['Materia'].to_list()
    newList=[]
    for item in list:                
        newList.append(normalize_string(item))                             
    dataframe['Materia']= newList 

    #extragem notele la practica de specialitate in functie de semestru
    global note_practica
    list_practica = csv_df[(csv_df['Materia']=='Practica de specialitate')]
    list_semestrul_ii = list_practica[list_practica['textbox2'].str.startswith('SEMESTRUL II')]['Nota'].tolist()
    list_semestrul_iv = list_practica[list_practica['textbox2'].str.startswith('SEMESTRUL IV')]['Nota'].tolist()
    note_practica = [max(list_semestrul_ii, default=''), max(list_semestrul_iv, default='')]
    
    # eliminam duplicatele din lista de materii pastrand-o doar pe cea cu nota cea mai mare
    dataframe.sort_values(['Materia','Nota'],ascending=[True,False], inplace=True)
    dataframe.drop_duplicates(subset=['Materia'], keep='first', inplace=True)
    
    # inlocuim nota 11 folosita pentru a putea face calcule pe coloana Nota cu calificativul admis "Adm"
    # dataframe['Nota'].replace(11,"Adm", inplace=True) 


# genereare pv echivalare scriind datele obtinute din situatia scolara .csv direct in macheta procesului verbal de echivalare
def prelucrare_date():

    # cream un dictionar cu datele din fisierul csv care are ca si chei denumirile materiilor si valori notele
    list1 = csv_df['Materia'].to_list()
    list2 = csv_df['Nota'].to_list()        
    new_dict ={}        
    for i in range (0,len(list1)):
        new_dict[list1[i]]=list2[i] 
    # date referitoare la exmatriculare sau intrerupere studii
    message = get_exmatriculare_intrerupere()
    #scriem notele si mentiuni asupra examenelor direct in macheta pv echivalare
    A = load_workbook(pv_path)
    A1 =load_workbook(catalog_path)
    B = A['Sheet1']
    B1 = A1['Sheet1']
    B.cell(row=8,column=1,value=student_name)
    B.cell(row=9,column=1,value=f'număr matricol: {nr_matricol}')
    B.cell(row=11,column=1,value=message)
    B1.cell(row=14,column=1,value=f'pentru studentul {student_name}')
    B1.cell(row=15,column=1,value=f'numar matricol: {nr_matricol} inmatriculat in anul III de studii')
    rand = 15    
    total_puncte=0
    while rand<B.max_row:                            
        if normalize_string(B.cell(row=rand, column=2).value) in list1:            
            B.cell(row=rand, column=4, value = list2[list1.index(normalize_string(B.cell(row=rand, column=2).value))])
            if normalize_string(B.cell(row=rand, column=2).value)=='Practica de specialitate':#introducere note la Practica de specialitate functie de semestru (cea mai mare nota corespunzatoare semestrului 2 respectiv 4)  
                B.cell(row=rand, column=4, value = note_practica[0])
                note_practica.pop(0)                
            if list2[list1.index(normalize_string(B.cell(row=rand, column=2).value))] > 4:
                B.cell(row=rand, column=5, value = B.cell(row=rand, column=2).value)
            else:
                B.cell(row=rand, column=5, value = 'Disciplina nepromovata')
            if normalize_string(B.cell(row=rand, column=2).value) == 'Practica de specialitate':
                if B.cell(row=rand, column=4).value == '':
                    B.cell(row=rand, column=5, value = 'Examen de diferenta')
                elif B.cell(row=rand, column=4).value < 5:
                    B.cell(row=rand, column=5, value = 'Disciplina nepromovata')
            #ignoram disciplina educatie fizica datorita particularitatilor ce trebuiesc analizate separat de operator
            if normalize_string(B.cell(row=rand, column=2).value) in ['Educatia fizica','Educatie fizica']:
                B.cell(row=rand, column=5, value = '')
                B.cell(row=rand, column=4, value = '')            
        else:
            if type(B.cell(row=rand, column=3).value) is int:
                B.cell(row=rand, column=5, value = 'Examen de diferenta')
                if B.cell(row=rand,column=2).value.startswith('Limba'):
                    B.row_dimensions[rand].hidden = True
        if type(B.cell(row=rand, column=4).value) is int:            
            if B.cell(row=rand,column=4).value > 4:
                total_puncte+=B.cell(row=rand,column=3).value*B.cell(row=rand,column=4).value                
        if normalize_string(B.cell(row=rand,column=2).value) == 'Total puncte':            
            B.cell(row=rand,column=4, value = total_puncte)
            total_puncte = 0
        rand+=1
    
    # calcul taxa de scolarizare
    rand =15
    taxa=0 
    
    while rand < B.max_row:           
        if normalize_string(B.cell(row=rand,column=2).value) not in ['Educatia fizica','Educatie fizica']:            
            if type(B.cell(row=rand,column=4).value) is int:
                if B.cell(row=rand,column=4).value < 5:
                    taxa += B.cell(row=rand,column=3).value*55                                       
            elif type(B.cell(row=rand, column=3).value) is int:                
                if B.cell(row=rand, column=4).value == '' or B.cell(row=rand,column=4).value ==None:
                    if B.row_dimensions[rand].hidden == False:
                        taxa += B.cell(row=rand,column=3).value*55                                                  
            if type(B.cell(row=rand,column=2).value) is str:
                if normalize_string(B.cell(row =rand, column=2).value).startswith('Optional'):
                    text = B.cell(row =rand, column=2).value.split('(',1)[1].split('din')
                    necesare = int(text[0].strip())
                    totale = int(text[1].split(')')[0].strip())                
                    contor=0
                    inner_rand = rand
                    while inner_rand <= rand + totale:
                        if type(B.cell(row=inner_rand,column=4).value) is int:
                            if B.cell(row=inner_rand,column=4).value > 4:
                                contor+=1                                               
                        inner_rand+=1
                    taxa +=(necesare-contor)*55*B.cell(row=rand+1,column=3).value
                    rand=inner_rand            
            if normalize_string(B.cell(row=rand,column=2).value) == 'Total puncte':                    
                    B.cell(row=rand,column=5, value = f'Taxa de plata : {taxa} lei')
                    taxa = 0        
        rand+=1
    rand =15
    rand1=19   
    while rand < B.max_row:
        if normalize_string(B.cell(row=rand,column=5).value) == 'Examen de diferenta' and B.row_dimensions[rand].hidden == False:            
            B1.cell(row=rand1, column=2, value=B.cell(row=rand,column=2).value)
            rand1+=2
        rand+=1
    output_folder = os.path.join(os.path.expanduser('~'), 'Desktop', 'Fisiere create pentru echivalare', f'{student_name}')
    os.makedirs(output_folder, exist_ok=True)
    output_file_path_pv = os.path.join(output_folder, f'PV_echivalare_{student_name}.xlsx')
    output_file_path_catalog = os.path.join(output_folder, f'Catalog_diferente_{student_name}.xlsx')
    A.save(output_file_path_pv)
    A1.save(output_file_path_catalog)

    # generam  ordonarea alfabetica a examenelor sutinute, si le scriem intr-un sitem de foldere creat pe desktop
    output_file_path_examene = os.path.join(output_folder, f'Examene_sustinute_ordine_alfabetica_{student_name}.xlsx')
    csv_df[['Materia','Nota']].to_excel(output_file_path_examene, index=False)  
     

# generare mesaj de finalizare si optiuni de continuare sau inchidere
def show_message():
    message = f"Au fost generate fisierele 'PV_echivalare_{student_name}' si 'Examene_sustinute_ordine_alfabetica_{student_name}' in folderul 'Fisiere create pentru echivalare'.\nContinuati cu alt student?"
    choice = messagebox.askyesno("Operatiunea s-a incheiat cu succes!", message)    
    if not choice:
        root.destroy()  # Close the application

# functia principala care ruleaza aplicatia
def rulare_program():
    # try:
        prelucrare_csv(csv_df)
        prelucrare_date()        
        show_message()
    # except Exception:        
        # messagebox.showerror("Eroare","S-a produs o eroare. Posibile cauze:\n- nu ati introdus numele studentului\n-nu ati selectat fisierele necesare\n-ati selectat fisiere gresite\nVerificati si incercati din nou!")
    
# Create the main application window
root = tk.Tk()
root.title("Generare proces verbal echivalare")

# Set the window size and position it in the center of the screen
window_width = 900
window_height = 500

screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

x = (screen_width // 2) - (window_width // 2)
y = (screen_height // 2) - (window_height // 2)

root.geometry(f"{window_width}x{window_height}+{x}+{y}")

# Set custom colors for the GUI
root.configure(bg='#344955')  # Grey background color

# Define function to create 'primary' style button with color that complements grey
def create_primary_button(text, command):
    button = tk.Button(root, text=text, command=command, font=('Helvetica', 15), padx=10, pady=5, bg='#50727B', fg='white', relief=tk.RAISED, cursor='hand2')
    return button

# creare butoane de selectare a fiseruluicu situatia scolara si a machetelor excel (echivalare si catalog)
csv_button = create_primary_button("Selecteaza fisierul care contine situatia scolara a studentului in format .csv", select_csv_file)
csv_button.pack(side=tk.TOP, pady=(100, 20))

excel_button =create_primary_button("Selecteaza fisierul care contine macheta procesului verbal de echivalare in format excel", select_excel_file)
excel_button.pack(side=tk.TOP, pady=20)

catalog_button =create_primary_button("Selecteaza fisierul care contine macheta catalogului de diferente in format excel", select_catalog_diferente)   
catalog_button.pack(side=tk.TOP, pady=20)

# creare buton de generare a fisierelor dorite
print_button =create_primary_button("Genereaza pv echivalare si catalog", rulare_program)
print_button.pack(side=tk.BOTTOM, pady=20)

# Run the application
root.mainloop()

